from __future__ import annotations

import argparse
import re
import sys
import traceback
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from sqlalchemy import or_, select, text
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.database import Base, SessionLocal, engine
from app.models import LessonRecord, Payment, Student, StudentSubject
from app.scripts.patch_db_schema import patch_student_subjects, patch_students
from app.services.class_types import (
    CLASS_TYPES,
    normalize_class_type,
)


DEFAULT_GRADE = "未分配"
DEFAULT_SUBJECT_NAME = "综合"
EXPECTED_DATABASE = "jinbi_backend_db"
REQUIRED_HEADERS = ["年级", "姓名", "学校", "班型", "电话"]
GRADE_ORDER = ["初一", "初二", "初三", "高一", "高二", "高三", DEFAULT_GRADE]


@dataclass
class ExcelStudentRow:
    row_number: int
    grade: str
    name: str
    school: str | None
    class_type: str
    phone: str | None


@dataclass
class DatabaseInfo:
    host: str
    configured_database: str
    current_database: str


@dataclass
class VerificationStats:
    students_count: int
    lesson_records_count: int
    payments_count: int
    student_subjects_count: int
    grade_class_rows: list[tuple[str, str, int]]


@dataclass
class ImportStats:
    excel_total: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    empty_phone: int = 0
    subject_placeholders_created: int = 0
    before_students: int = 0
    after_students: int = 0
    db_info: DatabaseInfo | None = None
    verification: VerificationStats | None = None
    reset_auto_increment_to: int | None = None
    class_type_counter: Counter[str] | None = None
    grade_counter: Counter[str] | None = None

    def __post_init__(self) -> None:
        self.class_type_counter = Counter()
        self.grade_counter = Counter()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return str(int(value))
        return format(value.normalize(), "f")
    return str(value).strip()


def normalize_phone(value: Any) -> str | None:
    text_value = normalize_text(value)
    if not text_value:
        return None

    compact = re.sub(r"\s+", "", text_value)
    numeric_candidate = compact
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", numeric_candidate):
        try:
            decimal_value = Decimal(numeric_candidate)
            if decimal_value == decimal_value.to_integral_value():
                return format(decimal_value.quantize(Decimal("1")), "f")
            return format(decimal_value.normalize(), "f").rstrip("0").rstrip(".")
        except (InvalidOperation, ValueError):
            return compact

    return compact


def normalize_grade(value: Any) -> str:
    grade = normalize_text(value)
    return grade or DEFAULT_GRADE


def normalize_school(value: Any) -> str | None:
    school = normalize_text(value)
    return school or None


def normalize_header(value: Any) -> str:
    return normalize_text(value).replace(" ", "")


def column_index(cell_ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", (cell_ref or "").upper())
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return max(index - 1, 0)


def read_excel_rows(file_path: Path) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    except ImportError:
        return read_excel_rows_stdlib(file_path)


def read_excel_rows_stdlib(file_path: Path) -> list[list[str]]:
    namespaces = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    with zipfile.ZipFile(file_path) as archive:
        shared_strings = read_shared_strings(archive, namespaces)
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rel_root.findall("pkgrel:Relationship", namespaces)
        }

        selected_sheet = None
        for sheet in workbook_root.findall("main:sheets/main:sheet", namespaces):
            if sheet.attrib.get("name") == "Sheet1":
                selected_sheet = sheet
                break
        if selected_sheet is None:
            selected_sheet = workbook_root.find("main:sheets/main:sheet", namespaces)
        if selected_sheet is None:
            raise ValueError("Excel 中没有可读取的工作表")

        relation_id = selected_sheet.attrib.get(f"{{{namespaces['rel']}}}id")
        target = rel_targets.get(relation_id or "")
        if not target:
            raise ValueError("无法定位 Sheet1 工作表数据")

        target_path = target.lstrip("/")
        if not target_path.startswith("xl/"):
            target_path = "xl/" + target_path

        worksheet_root = ET.fromstring(archive.read(target_path))
        rows: list[list[str]] = []
        for row in worksheet_root.findall("main:sheetData/main:row", namespaces):
            values: dict[int, str] = {}
            max_index = -1
            for cell in row.findall("main:c", namespaces):
                index = column_index(cell.attrib.get("r", ""))
                max_index = max(max_index, index)
                values[index] = read_cell_value(cell, shared_strings, namespaces)
            rows.append([values.get(index, "") for index in range(max_index + 1)])
        return rows


def read_shared_strings(archive: zipfile.ZipFile, namespaces: dict[str, str]) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []

    values: list[str] = []
    for item in root.findall("main:si", namespaces):
        parts = [node.text or "" for node in item.findall(".//main:t", namespaces)]
        values.append("".join(parts))
    return values


def read_cell_value(cell: ET.Element, shared_strings: list[str], namespaces: dict[str, str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        parts = [node.text or "" for node in cell.findall(".//main:t", namespaces)]
        return "".join(parts)

    value_node = cell.find("main:v", namespaces)
    if value_node is None or value_node.text is None:
        return ""

    raw_value = value_node.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)]
        except (IndexError, ValueError):
            return ""
    if cell_type == "b":
        return "1" if raw_value == "1" else "0"
    return raw_value


def extract_students_from_excel(file_path: Path) -> tuple[list[ExcelStudentRow], int, int]:
    rows = read_excel_rows(file_path)
    header_index = -1
    header_map: dict[str, int] = {}

    for index, row in enumerate(rows):
        normalized_cells = [normalize_header(value) for value in row]
        if all(header in normalized_cells for header in REQUIRED_HEADERS):
            header_index = index
            header_map = {header: normalized_cells.index(header) for header in REQUIRED_HEADERS}
            break

    if header_index < 0:
        raise ValueError("未找到表头：年级 | 姓名 | 学校 | 班型 | 电话")

    students: list[ExcelStudentRow] = []
    skipped = 0
    excel_total = 0

    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = {
            header: row[column] if column < len(row) else None
            for header, column in header_map.items()
        }
        if not any(normalize_text(values[header]) for header in REQUIRED_HEADERS):
            continue

        excel_total += 1
        name = normalize_text(values["姓名"])
        raw_class_type = normalize_text(values["班型"])
        class_type = normalize_class_type(raw_class_type)
        if not name:
            skipped += 1
            print(f"[skip] 第 {row_index} 行缺少姓名")
            continue
        if class_type not in CLASS_TYPES:
            skipped += 1
            print(f"[skip] 第 {row_index} 行班型无法识别：{raw_class_type or '空'}")
            continue

        students.append(
            ExcelStudentRow(
                row_number=row_index,
                grade=normalize_grade(values["年级"]),
                name=name,
                school=normalize_school(values["学校"]),
                class_type=class_type,
                phone=normalize_phone(values["电话"]),
            )
        )

    return students, excel_total, skipped


def column_nullable(db: Session, table_name: str, column_name: str) -> bool:
    nullable = db.execute(
        text(
            """
            SELECT IS_NULLABLE
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = :table_name
              AND COLUMN_NAME = :column_name
            """
        ),
        {"table_name": table_name, "column_name": column_name},
    ).scalar_one_or_none()
    return str(nullable or "NO").upper() == "YES"


def ensure_import_schema() -> None:
    Base.metadata.create_all(bind=engine)
    with engine.begin() as conn:
        patch_students(conn)
        patch_student_subjects(conn)


def generate_student_no(db: Session) -> str:
    year_prefix = f"S{datetime.now().year}"
    existing_numbers = db.scalars(
        select(Student.student_no).where(Student.student_no.like(f"{year_prefix}%"))
    ).all()

    max_sequence = 0
    for student_no in existing_numbers:
        suffix = str(student_no or "")[len(year_prefix) :]
        if suffix.isdigit():
            max_sequence = max(max_sequence, int(suffix))

    sequence = max_sequence + 1
    while True:
        candidate = f"{year_prefix}{sequence:04d}"
        if db.scalar(select(Student.id).where(Student.student_no == candidate)) is None:
            return candidate
        sequence += 1


def placeholder_phone(student_no: str) -> str:
    return f"NO_PHONE_{student_no}"


def resolve_phone(phone: str | None, student_no: str, phone_nullable: bool, existing_student: Student | None) -> str | None:
    if phone:
        return phone
    if phone_nullable:
        return None

    existing_phone = (existing_student.phone if existing_student is not None else "") or ""
    if existing_phone and not existing_phone.startswith("NO_PHONE_"):
        return existing_phone
    return placeholder_phone(student_no)


def find_existing_student(db: Session, row: ExcelStudentRow) -> Student | None:
    if row.phone:
        return db.scalar(select(Student).where(Student.phone == row.phone))

    stmt = select(Student).where(
        Student.name == row.name,
        Student.grade == row.grade,
    )
    if row.school:
        stmt = stmt.where(Student.school == row.school)
    else:
        stmt = stmt.where(or_(Student.school.is_(None), Student.school == ""))
    return db.scalar(stmt)


def ensure_default_subject(db: Session, student: Student) -> bool:
    subject = db.scalar(
        select(StudentSubject).where(
            StudentSubject.student_id == student.id,
            StudentSubject.subject_name == DEFAULT_SUBJECT_NAME,
        )
    )
    if subject is not None:
        return False

    db.add(
        StudentSubject(
            student_id=student.id,
            subject_name=DEFAULT_SUBJECT_NAME,
            total_hours=Decimal("0"),
            remaining_hours=Decimal("0"),
            deducted_hours=Decimal("0"),
        )
    )
    return True


def scalar_int(db: Session, sql: str) -> int:
    return int(db.execute(text(sql)).scalar_one() or 0)


def get_database_info(db: Session) -> DatabaseInfo:
    return DatabaseInfo(
        host=f"{settings.mysql_host}:{settings.mysql_port}",
        configured_database=settings.mysql_db,
        current_database=str(db.execute(text("SELECT DATABASE()")).scalar_one() or ""),
    )


def validate_database(info: DatabaseInfo) -> None:
    if info.current_database.lower() != EXPECTED_DATABASE.lower():
        raise RuntimeError(
            "导入脚本连接的数据库不是 "
            f"{EXPECTED_DATABASE}，当前 DATABASE()={info.current_database}，"
            f"配置 MYSQL_DB={info.configured_database}"
        )


def collect_verification(db: Session) -> VerificationStats:
    grade_class_rows = []
    for row in db.execute(
        text(
            """
            SELECT
              COALESCE(NULLIF(grade, ''), :default_grade) AS grade,
              COALESCE(NULLIF(class_type, ''), '未设置班型') AS class_type,
              COUNT(*) AS count
            FROM students
            WHERE isDeleted = 0 OR isDeleted IS NULL
            GROUP BY grade, class_type
            ORDER BY grade, class_type
            """
        ),
        {"default_grade": DEFAULT_GRADE},
    ):
        mapping = row._mapping
        grade_class_rows.append(
            (
                str(mapping["grade"] or DEFAULT_GRADE),
                str(mapping["class_type"] or "未设置班型"),
                int(mapping["count"] or 0),
            )
        )

    return VerificationStats(
        students_count=scalar_int(db, "SELECT COUNT(*) FROM students"),
        lesson_records_count=scalar_int(db, "SELECT COUNT(*) FROM lesson_records"),
        payments_count=scalar_int(db, "SELECT COUNT(*) FROM payments"),
        student_subjects_count=scalar_int(db, "SELECT COUNT(*) FROM student_subjects"),
        grade_class_rows=grade_class_rows,
    )


def collect_verification_after_commit() -> VerificationStats:
    db = SessionLocal()
    try:
        info = get_database_info(db)
        validate_database(info)
        return collect_verification(db)
    finally:
        db.close()


def clear_student_related_tables(db: Session) -> None:
    db.execute(text("DELETE FROM lesson_records"))
    db.execute(text("DELETE FROM payments"))
    db.execute(text("DELETE FROM student_subjects"))
    db.execute(text("DELETE FROM students"))


def reset_students_auto_increment(next_id: int) -> None:
    next_auto_increment = max(int(next_id), 1)
    with engine.begin() as conn:
        current_database = str(conn.execute(text("SELECT DATABASE()")).scalar_one() or "")
        if current_database.lower() != EXPECTED_DATABASE.lower():
            raise RuntimeError(
                f"重置 students 自增前发现数据库不正确：{current_database}"
            )
        conn.execute(text(f"ALTER TABLE students AUTO_INCREMENT = {next_auto_increment}"))


def import_students(rows: list[ExcelStudentRow], replace: bool) -> ImportStats:
    stats = ImportStats()
    stats.excel_total = len(rows)
    db = SessionLocal()
    next_student_id = 1

    try:
        with db.begin():
            stats.db_info = get_database_info(db)
            validate_database(stats.db_info)
            stats.before_students = scalar_int(db, "SELECT COUNT(*) FROM students")
            phone_nullable = column_nullable(db, "students", "phone")
            if replace:
                clear_student_related_tables(db)

            for row in rows:
                if row.phone is None:
                    stats.empty_phone += 1

                student = find_existing_student(db, row) if not replace else None
                if student is None:
                    student_no = generate_student_no(db)
                    student_kwargs: dict[str, Any] = {}
                    if replace:
                        student_kwargs["id"] = next_student_id
                        next_student_id += 1
                    student = Student(
                        **student_kwargs,
                        name=row.name,
                        phone=resolve_phone(row.phone, student_no, phone_nullable, None),
                        student_no=student_no,
                        grade=row.grade,
                        school=row.school,
                        avatar="",
                        class_type=row.class_type,
                        status="normal",
                        is_deleted=False,
                        deleted_at=None,
                    )
                    db.add(student)
                    db.flush()
                    stats.inserted += 1
                else:
                    student.name = row.name
                    student.phone = resolve_phone(row.phone, student.student_no, phone_nullable, student)
                    student.grade = row.grade
                    student.school = row.school
                    student.class_type = row.class_type
                    student.status = "normal"
                    student.is_deleted = False
                    student.deleted_at = None
                    stats.updated += 1

                if ensure_default_subject(db, student):
                    stats.subject_placeholders_created += 1

                stats.class_type_counter[row.class_type] += 1
                stats.grade_counter[row.grade] += 1

            db.flush()
            stats.after_students = scalar_int(db, "SELECT COUNT(*) FROM students")
            stats.verification = collect_verification(db)

        if replace:
            reset_students_auto_increment(next_student_id)
            stats.reset_auto_increment_to = max(next_student_id, 1)
            stats.verification = collect_verification_after_commit()
            stats.after_students = stats.verification.students_count

        return stats
    except Exception:
        db.rollback()
        traceback.print_exc()
        raise
    finally:
        db.close()


def print_counter(title: str, counter: Counter[str], preferred_order: list[str]) -> None:
    print(f"{title}：")
    printed = set()
    for key in preferred_order:
        if counter.get(key, 0):
            print(f"- {key}：{counter[key]}")
            printed.add(key)
    for key in sorted(counter.keys()):
        if key not in printed:
            print(f"- {key}：{counter[key]}")


def print_summary(stats: ImportStats, replace: bool) -> None:
    if stats.db_info is not None:
        print(f"数据库 host：{stats.db_info.host}")
        print(f"配置数据库名：{stats.db_info.configured_database}")
        print(f"当前数据库：{stats.db_info.current_database}")
    print(f"导入前 students 数量：{stats.before_students}")
    print(f"读取 Excel：{stats.excel_total} 人")
    print(f"replace 模式：{'是' if replace else '否'}")
    if replace:
        print("已清空旧学生数据")
        if stats.reset_auto_increment_to is not None:
            print(f"students AUTO_INCREMENT 已重置为：{stats.reset_auto_increment_to}")
    print(f"成功导入：{stats.inserted} 人")
    print(f"更新人数：{stats.updated} 人")
    print(f"跳过人数：{stats.skipped} 人")
    print(f"手机号为空：{stats.empty_phone} 人")
    print(f"导入后 students 数量：{stats.after_students}")
    print()
    print_counter("班型统计", stats.class_type_counter or Counter(), CLASS_TYPES)
    print()
    print_counter("年级统计", stats.grade_counter or Counter(), GRADE_ORDER)
    print()
    print("课时处理：")
    print(f"- 已创建 student_subjects 占位记录：{stats.subject_placeholders_created} 条")
    print("- totalHours = 0")
    print("- remainingHours = 0")
    print("- deductedHours = 0")
    print("- 未创建 lesson_records")
    print("- 未创建 payments")
    print()
    print("导入后数据库验证：")
    if stats.verification is not None:
        print(f"SELECT COUNT(*) FROM students; => {stats.verification.students_count}")
        print("SELECT grade, class_type, COUNT(*) ... =>")
        for grade, class_type, count in stats.verification.grade_class_rows:
            print(f"- {grade} / {class_type}：{count}")
        print(f"SELECT COUNT(*) FROM lesson_records; => {stats.verification.lesson_records_count}")
        print(f"SELECT COUNT(*) FROM payments; => {stats.verification.payments_count}")
        print(f"SELECT COUNT(*) FROM student_subjects; => {stats.verification.student_subjects_count}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从 Excel 导入学员基础信息")
    parser.add_argument("--file", required=True, help="Excel 文件路径")
    parser.add_argument("--replace", action="store_true", help="清空学生相关测试数据后重新导入")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    file_path = Path(args.file).expanduser()
    if not file_path.is_absolute():
        file_path = Path.cwd() / file_path
    file_path = file_path.resolve(strict=False)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{file_path}")

    print(f"Excel 文件：{file_path}")
    rows, excel_total, skipped = extract_students_from_excel(file_path)
    ensure_import_schema()
    stats = import_students(rows, replace=bool(args.replace))
    stats.excel_total = excel_total
    stats.skipped += skipped
    print_summary(stats, replace=bool(args.replace))


if __name__ == "__main__":
    main()
